import pandas as pd
from pathlib import Path
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import os

def try_parse_float(x: object) -> object:
    s = str(x).strip().replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return x

def PersMargem(caminhoBase,nameMargem:str):
    caminho_base = Path(caminhoBase)
    arquivo = Path(nameMargem)

    # 1) Carrega modelo para cabeçalho e estilos
    wb_modelo = load_workbook(os.path.join(caminho_base, 'modelo.xlsx'), data_only=False)
    ws_modelo = wb_modelo['modelo']
    header = [cell.value for cell in ws_modelo[1]]
    cols_excel = ['B','D','H','I','J','K','L','N','S','Y','Z','AA','AB','AC','AD','AE']
    dfs ={}

    # 2) Lê abas em DataFrames
    all_sheets = pd.read_excel(arquivo, sheet_name=None, header=0, engine='openpyxl')
    for nome_abas, df in all_sheets.items():
        df.columns = header
        df.iloc[1:].reset_index(drop=True)
        dfs[nome_abas] = df
    
    # 3) Limpeza de colunas numéricas (letras do Excel → índices pandas)
    idxs = [column_index_from_string(c)-1 for c in cols_excel]
    for df in all_sheets.values():
        for idx in idxs:
            col = df.columns[idx]
            df[col] = df[col].apply(try_parse_float)

    # 4) Sobrescreve abas no arquivo
    with pd.ExcelWriter(arquivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        for sheet_mane, df in all_sheets.items():
            df.to_excel(writer, sheet_name=sheet_mane, index=False, header=header)

    # 5) Copia estilos do cabeçalho do modelo e ajusta larguras num único loop
    wb_dest = load_workbook(arquivo)
    for sheet_name in all_sheets.keys():
        ws_dest = wb_dest[sheet_name]
        # ws_modelo[1] é a primeira linha do modelo; ws_dest.columns são as colunas do destino
        for col_idx, (cell_model, col_cells) in enumerate(zip(ws_modelo[1], ws_dest.columns), start=1):
            # copia estilo do cabeçalho
            dest_cell = ws_dest.cell(row=1, column=col_idx)
            dest_cell.fill          = copy(cell_model.fill)
            dest_cell.font          = copy(cell_model.font)
            dest_cell.border        = copy(cell_model.border)
            dest_cell.alignment     = copy(cell_model.alignment)
            dest_cell.number_format = cell_model.number_format

            # ajusta largura dessa coluna
            letter = col_cells[0].column_letter
            max_len = max(len(str(c.value or "")) for c in col_cells)
            ws_dest.column_dimensions[letter].width = max_len + 2

    wb_dest.save(arquivo)